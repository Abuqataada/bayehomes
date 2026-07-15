import csv
import io
import uuid
from datetime import date, datetime, timedelta
from flask import Blueprint, render_template, redirect, url_for, flash, request, jsonify
from flask_login import login_required, current_user
from sqlalchemy import func, or_, and_
from models import (
    db, User, StaffProfile, ActivityLog, Lead, Sale, Property, LeadStatus,
    UserRole, StaffDailyReport, StaffAttendanceRecord, StaffDevice, StaffRating,
    StaffChatMessage, StaffNotification, StaffNotificationRead
)
from forms.staff_forms import StaffRegistrationForm, StaffEditForm, StaffFilterForm
from decorators import role_required

staff_bp = Blueprint('staff', __name__, url_prefix='/staff')


def role_value(user):
    return getattr(user.role, 'value', user.role)


def staff_query():
    return User.query.filter(User.role.in_([UserRole.ADMIN, UserRole.STAFF]))


def log_activity(user_id, action, details=None, request_obj=None):
    ip = request_obj.remote_addr if request_obj else None
    log = ActivityLog(user_id=user_id, action=action, details=details or {}, ip_address=ip)
    db.session.add(log)
    db.session.commit()


def parse_date(value):
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if not value:
        return date.today()
    text = str(value).strip()
    for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y', '%d-%m-%Y'):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            pass
    return date.today()


@staff_bp.route('/dashboard')
@login_required
@role_required(['admin', 'staff'])
def dashboard():
    if current_user.is_admin_user():
        total_staff = staff_query().count()
        active_staff = staff_query().filter(User.is_active == True).count()
        total_leads = Lead.query.count()
        converted_leads = Lead.query.filter(Lead.status == LeadStatus.CONVERTED).count()
        total_sales = Sale.query.count()
        revenue = db.session.query(func.sum(Sale.total_amount)).scalar() or 0
        reports_today = StaffDailyReport.query.filter_by(report_date=date.today()).count()
        unread_messages = StaffChatMessage.query.filter_by(recipient_id=current_user.id, is_read=False).count()
        recent_activities = ActivityLog.query.order_by(ActivityLog.created_at.desc()).limit(10).all()
        notifications = StaffNotification.query.order_by(StaffNotification.created_at.desc()).limit(5).all()
        return render_template('staff/dashboard.html', total_staff=total_staff, active_staff=active_staff,
                               total_leads=total_leads, converted_leads=converted_leads, total_sales=total_sales,
                               revenue=revenue, reports_today=reports_today, unread_messages=unread_messages,
                               recent_activities=recent_activities, notifications=notifications)

    leads_assigned = Lead.query.filter_by(assigned_to=current_user.id).count()
    leads_converted = Lead.query.filter_by(assigned_to=current_user.id, status=LeadStatus.CONVERTED).count()
    sales_closed = Sale.query.filter_by(staff_id=current_user.id, status='completed').count()
    managed_properties = Property.query.filter_by(assigned_staff_id=current_user.id).all()
    latest_notifications = StaffNotification.query.order_by(StaffNotification.created_at.desc()).limit(5).all()
    reports = StaffDailyReport.query.filter_by(staff_id=current_user.id).order_by(StaffDailyReport.report_date.desc()).limit(7).all()
    return render_template('staff/dashboard.html', staff=current_user, leads_assigned=leads_assigned,
                           leads_converted=leads_converted, sales_closed=sales_closed,
                           managed_properties=managed_properties, notifications=latest_notifications, reports=reports)


@staff_bp.route('/')
@login_required
@role_required(['admin'])
def list_staff():
    form = StaffFilterForm(request.args, meta={'csrf': False})
    query = staff_query()
    if form.search.data:
        search = f"%{form.search.data}%"
        # Ensure we can search by staff_id (StaffProfile) along with name/email
        query = query.join(
            StaffProfile,
            StaffProfile.user_id == User.id,
            isouter=True
        ).filter(
            or_(
                StaffProfile.staff_id.ilike(search),
                User.first_name.ilike(search),
                User.last_name.ilike(search),
                User.email.ilike(search),
            )
        )
    if form.department.data:
        query = query.join(
            StaffProfile,
            StaffProfile.user_id == User.id
        ).filter(StaffProfile.department == form.department.data)
    if form.is_active.data:
        query = query.filter(User.is_active == (form.is_active.data == 'true'))
    staff_members = query.order_by(User.created_at.desc()).all()
    return render_template('staff/list.html', staff_members=staff_members, form=form)


@staff_bp.route('/create', methods=['GET', 'POST'])
@login_required
@role_required(['admin'])
def create_staff():
    form = StaffRegistrationForm()
    if form.validate_on_submit():
        user = User(email=form.email.data, username=form.email.data.split('@')[0], first_name=form.first_name.data,
                    last_name=form.last_name.data, phone=form.phone.data, role=UserRole.STAFF,
                    is_active=form.is_active.data, email_verified=True)
        user.set_password(form.password.data)
        db.session.add(user)
        db.session.flush()
        profile = StaffProfile(user_id=user.id, staff_id=f"BH-{datetime.utcnow().year}-{uuid.uuid4().hex[:5].upper()}",
                               department=form.department.data, role=form.role.data,
                               reports_to=form.reports_to.data or None)
        db.session.add(profile)
        db.session.commit()
        log_activity(current_user.id, f"Created staff {user.email}", request_obj=request)
        flash('Staff created successfully.', 'success')
        return redirect(url_for('staff.list_staff'))
    return render_template('staff/form.html', form=form, title='Create Staff')


@staff_bp.route('/<staff_id>/edit', methods=['GET', 'POST'])
@login_required
@role_required(['admin'])
def edit_staff(staff_id):
    user = User.query.get_or_404(staff_id)
    form = StaffEditForm(obj=user)
    if user.staff_profile and request.method == 'GET':
        form.department.data = getattr(user.staff_profile.department, 'value', user.staff_profile.department)
        form.role.data = user.staff_profile.role
        form.reports_to.data = user.staff_profile.reports_to or ''
    if form.validate_on_submit():
        user.first_name = form.first_name.data
        user.last_name = form.last_name.data
        user.phone = form.phone.data
        user.is_active = form.is_active.data
        if not user.staff_profile:
            db.session.add(StaffProfile(user_id=user.id, staff_id=f"BH-{datetime.utcnow().year}-{uuid.uuid4().hex[:5].upper()}"))
            db.session.flush()
        user.staff_profile.department = form.department.data
        user.staff_profile.role = form.role.data
        user.staff_profile.reports_to = form.reports_to.data or None
        db.session.commit()
        flash('Staff updated.', 'success')
        return redirect(url_for('staff.list_staff'))
    return render_template('staff/form.html', form=form, title='Edit Staff', staff=user)


@staff_bp.route('/<staff_id>/profile')
@login_required
@role_required(['admin', 'staff'])
def view_profile(staff_id):
    user = User.query.get_or_404(staff_id)
    if current_user.id != user.id and not current_user.is_admin_user():
        flash('You do not have permission.', 'danger')
        return redirect(url_for('staff.dashboard'))
    ratings = StaffRating.query.filter_by(staff_id=user.id).order_by(StaffRating.created_at.desc()).limit(12).all()
    reports = StaffDailyReport.query.filter_by(staff_id=user.id).order_by(StaffDailyReport.report_date.desc()).limit(10).all()
    properties = Property.query.filter_by(assigned_staff_id=user.id).all()
    return render_template('staff/profile.html', staff=user, ratings=ratings, reports=reports, properties=properties)


@staff_bp.route('/reports', methods=['GET', 'POST'])
@login_required
@role_required(['admin', 'staff'])
def reports():
    if request.method == 'POST' and not current_user.is_admin_user():
        report = StaffDailyReport(
            staff_id=current_user.id,
            report_date=parse_date(request.form.get('report_date')),
            summary=request.form.get('summary', '').strip(),
            leads_contacted=int(request.form.get('leads_contacted') or 0),
            site_visits=int(request.form.get('site_visits') or 0),
            sales_closed=int(request.form.get('sales_closed') or 0),
            challenges=request.form.get('challenges')
        )
        if not report.summary:
            flash('Daily report summary is required.', 'danger')
        else:
            db.session.add(report)
            db.session.commit()
            flash('Daily report submitted.', 'success')
        return redirect(url_for('staff.reports'))
    query = StaffDailyReport.query
    if not current_user.is_admin_user():
        query = query.filter_by(staff_id=current_user.id)
    all_reports = query.order_by(StaffDailyReport.report_date.desc(), StaffDailyReport.created_at.desc()).all()
    return render_template('staff/reports.html', reports=all_reports)


@staff_bp.route('/chat', methods=['GET', 'POST'])
@login_required
@role_required(['admin', 'staff'])
def chat():
    selected_id = request.args.get('with')
    if request.method == 'POST':
        recipient_id = request.form.get('recipient_id')
        message = request.form.get('message', '').strip()
        if recipient_id and message:
            db.session.add(StaffChatMessage(sender_id=current_user.id, recipient_id=recipient_id, message=message))
            db.session.commit()
            flash('Message sent.', 'success')
            return redirect(url_for('staff.chat', **{'with': recipient_id}))
        flash('Select a recipient and enter a message.', 'danger')
    users = staff_query().filter(User.id != current_user.id).order_by(User.first_name).all()
    if not selected_id and users:
        selected_id = users[0].id
    messages = []
    if selected_id:
        messages = StaffChatMessage.query.filter(or_(
            and_(StaffChatMessage.sender_id == current_user.id, StaffChatMessage.recipient_id == selected_id),
            and_(StaffChatMessage.sender_id == selected_id, StaffChatMessage.recipient_id == current_user.id)
        )).order_by(StaffChatMessage.created_at.asc()).all()
        StaffChatMessage.query.filter_by(sender_id=selected_id, recipient_id=current_user.id, is_read=False).update({'is_read': True})
        db.session.commit()
    return render_template('staff/chat.html', users=users, selected_id=selected_id, messages=messages)


@staff_bp.route('/notifications', methods=['GET', 'POST'])
@login_required
@role_required(['admin', 'staff'])
def notifications():
    if request.method == 'POST' and current_user.is_admin_user():
        title = request.form.get('title', '').strip()
        message = request.form.get('message', '').strip()
        if title and message:
            db.session.add(StaffNotification(title=title, message=message, created_by=current_user.id))
            db.session.commit()
            flash('Notification sent to all staff.', 'success')
        else:
            flash('Title and message are required.', 'danger')
        return redirect(url_for('staff.notifications'))
    notifications = StaffNotification.query.order_by(StaffNotification.created_at.desc()).all()
    return render_template('staff/notifications.html', notifications=notifications)


@staff_bp.route('/notifications/<notification_id>/read', methods=['POST'])
@login_required
@role_required(['admin', 'staff'])
def mark_notification_read(notification_id):
    exists = StaffNotificationRead.query.filter_by(notification_id=notification_id, staff_id=current_user.id).first()
    if not exists:
        db.session.add(StaffNotificationRead(notification_id=notification_id, staff_id=current_user.id))
        db.session.commit()
    return redirect(url_for('staff.notifications'))


@staff_bp.route('/ratings')
@login_required
@role_required(['admin'])
def ratings():
    period = request.args.get('period', 'monthly')
    today = date.today()
    if period == 'yearly':
        start = date(today.year, 1, 1)
        label = str(today.year)
    else:
        start = date(today.year, today.month, 1)
        label = today.strftime('%Y-%m')
    staff_members = staff_query().filter(User.role == UserRole.STAFF).all()
    rows = []
    for staff in staff_members:
        attendance = StaffAttendanceRecord.query.filter(StaffAttendanceRecord.staff_id == staff.id, StaffAttendanceRecord.attendance_date >= start).count()
        reports_count = StaffDailyReport.query.filter(StaffDailyReport.staff_id == staff.id, StaffDailyReport.report_date >= start).count()
        sales_count = Sale.query.filter(Sale.staff_id == staff.id, Sale.sale_date >= datetime.combine(start, datetime.min.time())).count()
        attendance_score = min(attendance * 4, 40)
        reporting_score = min(reports_count * 3, 30)
        sales_score = min(sales_count * 10, 30)
        total = attendance_score + reporting_score + sales_score
        rating = StaffRating.query.filter_by(staff_id=staff.id, period_type=period, period_label=label).first()
        if not rating:
            rating = StaffRating(staff_id=staff.id, period_type=period, period_label=label)
            db.session.add(rating)
        rating.attendance_score = attendance_score
        rating.reporting_score = reporting_score
        rating.sales_score = sales_score
        rating.total_score = total
        rows.append((staff, rating, attendance, reports_count, sales_count))
    db.session.commit()
    return render_template('staff/ratings.html', rows=rows, period=period)


@staff_bp.route('/attendance/chart', methods=['GET'])
@login_required
@role_required(['admin'])
def attendance_chart():
    staff_members = staff_query().filter(User.role == UserRole.STAFF).order_by(User.first_name).all()

    range_days = request.args.get('range', 14, type=int)
    if range_days not in (14, 30, 90):
        range_days = 14

    selected_staff_id = request.args.get('staff_id')
    if not selected_staff_id and staff_members:
        selected_staff_id = staff_members[0].id

    selected_staff = None
    for s in staff_members:
        if s.id == selected_staff_id:
            selected_staff = s
            break

    if not selected_staff:
        return render_template(
            'staff/attendance_chart.html',
            staff_members=staff_members,
            selected_staff_id='',
            range_days=range_days,
            selected_staff_name='',
            labels=[],
            values=[]
        )

    end_date = date.today()
    start_date = end_date - timedelta(days=range_days - 1)

    # Aggregate counts per day
    rows = (
        StaffAttendanceRecord.query
        .with_entities(StaffAttendanceRecord.attendance_date, func.count(StaffAttendanceRecord.id))
        .filter(
            StaffAttendanceRecord.staff_id == selected_staff.id,
            StaffAttendanceRecord.attendance_date >= start_date,
            StaffAttendanceRecord.attendance_date <= end_date
        )
        .group_by(StaffAttendanceRecord.attendance_date)
        .order_by(StaffAttendanceRecord.attendance_date.asc())
        .all()
    )

    counts_by_date = {d: c for d, c in rows}

    labels = []
    values = []
    for i in range(range_days):
        day = start_date + timedelta(days=i)
        labels.append(day.strftime('%Y-%m-%d'))
        values.append(int(counts_by_date.get(day, 0)))

    return render_template(
        'staff/attendance_chart.html',
        staff_members=staff_members,
        selected_staff_id=selected_staff.id,
        range_days=range_days,
        selected_staff_name=selected_staff.get_full_name(),
        labels=labels,
        values=values
    )


@staff_bp.route('/attendance/import', methods=['GET', 'POST'])
@login_required
@role_required(['admin'])
def import_attendance():
    imported = 0

    def normalize_key(key: str) -> str:
        if key is None:
            return ''
        return str(key).strip().lower().replace(' ', '_').replace('-', '_')

    def normalize_row_keys(d: dict) -> dict:
        out = {}
        for k, v in (d or {}).items():
            out[normalize_key(k)] = v
        return out

    if request.method == 'POST':
        upload = request.files.get('attendance_file')
        if not upload or not upload.filename:
            flash('Upload an Excel or CSV attendance file.', 'danger')
            return redirect(url_for('staff.import_attendance'))

        filename = upload.filename
        rows = []

        if filename.lower().endswith('.csv'):
            stream = io.StringIO(upload.stream.read().decode('utf-8-sig'))
            reader = csv.DictReader(stream)
            rows = [normalize_row_keys(r) for r in reader]
        else:
            try:
                from openpyxl import load_workbook
                workbook = load_workbook(upload, data_only=True)
                sheet = workbook.active
                raw_headers = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
                headers = [normalize_key(h) for h in raw_headers]
                for raw in sheet.iter_rows(min_row=2, values_only=True):
                    rows.append(normalize_row_keys(dict(zip(headers, raw))))
            except Exception as exc:
                flash(f'Could not read Excel file: {exc}', 'danger')
                return redirect(url_for('staff.import_attendance'))

        for row in rows:
            device_id = row.get('device_id')
            name = row.get('name')  # fallback if device mapping isn't present

            user = None
            if device_id:
                mapping = StaffDevice.query.filter_by(device_id=str(device_id).strip()).first()
                user = mapping.staff if mapping else None

            # Optional fallback by name only if present on row
            if not user and name:
                name_parts = str(name).strip().split()
                if len(name_parts) >= 2:
                    first = name_parts[0]
                    last = ' '.join(name_parts[1:])
                    user = User.query.filter(User.first_name.ilike(f'%{first}%'), User.last_name.ilike(f'%{last}%')).first()
                else:
                    user = User.query.filter(
                        or_(User.first_name.ilike(f'%{name}%'), User.last_name.ilike(f'%{name}%'))
                    ).first()

            if not user:
                continue

            record = StaffAttendanceRecord(
                staff_id=user.id,
                attendance_date=parse_date(row.get('date') or row.get('attendance_date')),
                status=str(row.get('status') or 'present').lower(),
                check_in=str(row.get('time_in') or row.get('check_in') or ''),
                check_out=str(row.get('time_out') or row.get('check_out') or ''),
                source_file=filename
            )
            db.session.add(record)
            imported += 1

        db.session.commit()
        flash(f'Imported {imported} attendance records.', 'success')
        return redirect(url_for('staff.ratings'))

    return render_template('staff/import_attendance.html')


@staff_bp.route('/activity')
@login_required
@role_required(['admin'])
def activity_logs():
    logs = ActivityLog.query.order_by(ActivityLog.created_at.desc()).limit(200).all()
    return render_template('staff/activity_logs.html', logs=logs)


@staff_bp.route('/push/subscribe', methods=['POST'])
@login_required
@role_required(['admin', 'staff'])
def push_subscribe():
    return jsonify({'ok': True, 'message': 'Browser push hook received. Configure VAPID/web-push credentials in production to deliver OS-level notifications.'})
